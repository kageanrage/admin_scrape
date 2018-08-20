import logging
logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')
#logging.disable(logging.CRITICAL)     # switches off logging

logging.debug('Importing modules')
import bs4,re, openpyxl, os, sqlite3, requests, time, smtplib, pprint
from openpyxl.styles import Font, NamedStyle, PatternFill
from selenium import webdriver
from bs4 import BeautifulSoup
from config import Config   # this imports the config file where the private data sits

logging.debug('Start of program')
logging.debug('Checking if Laptop or Desktop (and opening relevant local HTML files if using test HTML)')

cfg = Config()      # create an instance of the Config class, essentially brings private config data into play

# changes logic depending on if I'm using Laptop or Desktop
# Example files - using saved HTML in 2 different directories. Toggle on for test mode, or off for live.

if os.getcwd() == cfg.laptop_dir:   #Using Laptop
    logging.debug('Laptop PC detected')
    localFile = open(cfg.laptop_localfile)
    exampleSoup = bs4.BeautifulSoup(localFile, "html.parser")  # turns the HTML into a beautiful soup object
    exampleNewHTMLFile = open(cfg.laptop_ex_html_file)
    exampleNewSoup = bs4.BeautifulSoup(exampleNewHTMLFile, "html.parser")  # turns the HTML into a beautiful soup object
    exampleOldHTMLFile = open(cfg.laptop_ex_old_html_file)
    exampleOldSoup = bs4.BeautifulSoup(exampleOldHTMLFile, "html.parser")  # turns the HTML into a beautiful soup object

elif os.getcwd() == cfg.desktop_dir:    #Using Desktop
    logging.debug('Desktop PC detected')
    localFile = open(cfg.desktop_localfile)
    exampleSoup = bs4.BeautifulSoup(localFile, "html.parser")  # turns the HTML into a beautiful soup object
    exampleNewHTMLFile = open(cfg.desktop_ex_html_file)
    exampleNewSoup = bs4.BeautifulSoup(exampleNewHTMLFile, "html.parser")  # turns the HTML into a beautiful soup object
    exampleOldHTMLFile = open(cfg.desktop_ex_old_html_file)
    exampleOldSoup = bs4.BeautifulSoup(exampleOldHTMLFile, "html.parser")  # turns the HTML into a beautiful soup object

def download_soup():
    chrome_path = r'C:\Program Files\Python37\chromedriver.exe'
    driver = webdriver.Chrome(chrome_path)
    driver.get(cfg.survey_admin_URL) # load survey admin page
    emailElem = driver.find_element_by_id('UserName') #enter username & password and submit
    emailElem.send_keys(cfg.uname)
    passElem = driver.find_element_by_id('Password')
    passElem.send_keys(cfg.pwd)
    passElem.submit()
    time.sleep(5)   # wait 5 seconds to log in
    content = driver.page_source
    soup = bs4.BeautifulSoup(content, "html.parser")
    # logging.debug('Newly downloaded soup looks like this:\n\n', soup)
    #soupFile = open(htmlFileName, "w")
    #soupFile.write(str(soup))
    #soupFile.close()
    return soup

def process_soup(soup):
    logging.debug('Starting table isolation')
    tableOnly = soup.select(
        'table')  # isolates the table (which is the only bit I need) from the HTML. Type is list, was expecting BS4 object
    #logging.debug('tableOnly looks like this:\n\n\n',tableOnly)
    logging.debug('Converting bs4 object into string')
    tableString = str(tableOnly)  # converts the bs4 object to a string
    #logging.debug('tableString looks like this:\n\n\n',tableString)
    # May not be able to isolate further within BS4 so switching to regex to parse.
    # TO DO: create a regex to identify each project on the Admin page
    logging.debug('Defining RegEx')
    # projectsRegex = re.compile('<a href="(.{80,105})">(.{3,50})<\/a><\/td><td class="clickable">(.{3,50})<\/td><td class="clickable">(.{3,10})<\/td><td class="clickable">(.{3,30})<\/td>(.{80,130})201\d<\/td><td class="clickable">(\d+)<\/td><td class="clickable">(\d+)<\/td><td class="clickable">(\d+)<\/td><td class="clickable">(\d+)<\/td><td class="clickable">(\d+)')
    #projectsRegex = re.compile(
    #    '<a href="(.{80,105})">(.{3,50})<\/a><\/td><td class="clickable">(.{3,50})<\/td><td class="clickable">(.{3,10})<\/td><td class="clickable">(.{3,30})<\/td>(.{80,130})201\d<\/td><td class="clickable">(\d+)<\/td><td class="clickable">(\d+)<\/td><td class="clickable">(\d+)<\/td><td class="clickable">(\d+)<\/td><td class="clickable">(\d+)<\/td><td class="published t-center clickable"><span class="((True)|(False))">((True)|(False))<\/span><\/td><\/tr><tr class="gridrow(_alternate)? selectable-row"><td class="clickable">')  # alternative Regex which incorporates 'True' or 'False' being on site
    projectsRegex = re.compile(
    '<a href="(.{10,105})">(.{3,50})<\/a><\/td><td class="clickable">(.{3,50})<\/td><td class="clickable">(.{3,10})<\/td><td class="clickable">(.{3,30})<\/td>(.{80,130})201\d<\/td><td class="clickable">(\d+)?<\/td><td class="clickable">(\d+)?<\/td><td class="clickable">(\d+)?<\/td><td class="clickable">(\d+)?<\/td><td class="clickable">(\d+)?<\/td><td class="published t-center clickable"><span class="((True)|(False))">((True)|(False))<\/span><\/td><\/tr><tr class="gridrow(_alternate)? selectable-row"><td class="clickable">') # 3rd iteration of regex
    # TO DO: Return all examples of regex findall search
    logging.debug('Conducting regex findall search')
    mo = projectsRegex.findall(tableString)
    #print('newly created mo looks like this:\n\n',mo)
    return mo

def listCreator(valueList):   #this function takes in a MO from the regex and creates and returns a per-project list, ordered as per the headings list below
    #headings = ['URL','Alias','Survey name','Project number','Client name','junk','Expected LOI','Actual LOI','Completes','Screen Outs','Quota Fulls','Live on site'] #here I've added 'Live on Site'
    newList = []
    #logging.debug('Start of list creation for',valueList[3])
    for i in range(0,12):
        newList.append(valueList[i])
    completes = int(valueList[8])
    QFs = int(valueList[10])
    SOs = int(valueList[9])
    if completes == 0 | SOs == 0 | QFs == 0:
        incidence = 0
        QFincidence = 0
    else:
        incidence = (completes / (completes + SOs))
        QFincidence = (completes / (completes + SOs + QFs))
    newList.append(incidence)
    newList.append(QFincidence)
    #logging.debug('newList is:',newList)
    #logging.debug('valueList is',valueList[0:12])
    #logging.debug('{} C / {} C + {} SOs + {} QFs = {} IR.'.format(completes,completes,SOs,QFs,incidence))
    #logging.debug(newDict)
    return newList









def create_masterList(mo):     #creates a list of all projects in given MO, first row will be headings
    #global masterList
    mList = [['URL', 'Alias', 'Survey name', 'Project number', 'Client name', 'junk', 'Expected LOI', 'Actual LOI',
                   'Completes', 'Screen Outs', 'Quota Fulls', 'Live on site', 'Incidence Rate', 'QF IR']]
    for i in range(0, len(mo) - 1):
        mList.append(listCreator(mo[i]))
    return mList

def create_topList(mo, num):    #num = how long you want the list to be
    tList = []
    # top10List = [['URL', 'Alias', 'Survey name', 'Project number', 'Client name', 'junk', 'Expected LOI', 'Actual LOI',
    #                'Completes', 'Screen Outs', 'Quota Fulls', 'Live on site', 'Incidence Rate', 'QF IR']]
    for i in range(0, num):
        tList.append(listCreator(mo[i]))
    return tList

def new_project_search(newList,oldList):

    matches = []
    unmatched = []

    for newProject in newList:
        unmatched.append(newProject[3])   #this should make a list with all the Project numbers in newList

    for newProject in newList:
        for oldProject in oldList:
            if newProject[3] == oldProject[3]:
                matches.append(newProject[3])
                #if newProject[3] not in unmatched:
                #    raise Exception('Project not found in unmatched list, cannot remove')
                try:
                    unmatched.remove(newProject[3]) #this should remove all matches so that unmatched is the list of non-matched jobs
                except:
                    print(newProject[3],'could not be removed')
                    pass

    #print('Unmatched are as follows: ',unmatched)
    print('List of matched items: ',matches)
    return(unmatched)

def excel_export(list):     #### THIS FUNCTION IS THE EXPORT TO EXCEL  #####
    logging.debug('Excel section - creating workbook object')
    wb = openpyxl.Workbook()  # create excel workbook object
    wb.save('admin.xlsx')  # save workbook as admin.xlsx
    sheet = wb.active  # create sheet object as the Active sheet from the workbook object
    wb.save('admin.xlsx')  # save workbook as admin.xlsx
    # LIST-BASED POPULATION OF EXCEL SHEET
    for row, rowData in enumerate(list,
                                  1):  # where row is a number starting with 1, increasing each loop, and rowData = each masterList item
        for column in range(1, 15):  # where column is a number starting with 1 and ending with 14
            cell = sheet.cell(row=row, column=column)  # so on first loop, row = 2, col = 1
            v = rowData[column - 1]
            try:
                v = float(v)  # try to convert value to a float, so it will store numbers as numbers and not strings
            except ValueError:
                pass  # if it's not a number and therefore returns an error, don't try to convert it to a number
            cell.value = v  # write the value (v) to the cell
            if (column == 13) | (column == 14):  # for all cells in column 13 or 14 (IR / QFIR)
                cell.style = 'Percent'  # ... change cell format (style) to 'Percent', a built-in style within openpyxl

    # this section populates the first row in the sheet (headings) with bold style
    #make_bold(sheet, wb, sheet['A1':'N1'])    #Calls the make_bold function on first row of excel sheet
    wb.save('admin.xlsx')  # save workbook as admin.xlsx
    logging.debug('Excel workbook completed and saved')

def excel_export_dict(dict, filename):     #### Modifying excel_export list fn to work with master_dict  #####
    logging.debug('Excel section - creating workbook object')
    wb = openpyxl.Workbook()  # create excel workbook object
    wb.save(filename)  # save workbook
    sheet = wb.active  # create sheet object as the Active sheet from the workbook object
    wb.save(filename)  # save workbook
    headingsList = ['URL','Alias','Survey name','Project number','Client name','junk','Expected LOI','Actual LOI','Completes','Screen Outs','Quota Fulls','Live on site', 'incidence', 'QFincidence']
    # DICT-BASED POPULATION OF EXCEL SHEET - NOT YET UPDATED BELOW THIS #####

    # this bit populates and emboldens the first row
    row = 1
    for column in range(0,len(headingsList)):
        cell = sheet.cell(row=row, column=column+1)
        cell.value = headingsList[column]
    make_bold(sheet, wb, sheet['A1':'N1'])    #Calls the make_bold function on first row of excel sheet

    #this bit then populates the rest of the sheet with the masterDict content
    for row, item_tuple in enumerate(dict.items(), 2):
        # print(f'row is {row}, key is {item_tuple[0]}, project dict is{item_tuple[1]}')
        for column, heading in enumerate(headingsList, 1):
            # print(f"row is {row}, column is {column} heading is {heading}, nested value is {item_tuple[1].get(heading)}")
            cell = sheet.cell(row=row, column=column)  # so on first loop, row = 2, col = 1
            v = item_tuple[1].get(heading)
            try:
                v = float(v)  # try to convert value to a float, so it will store numbers as numbers and not strings
            except ValueError:
                pass  # if it's not a number and therefore returns an error, don't try to convert it to a number
            cell.value = v
            if (column == 13) | (column == 14):  # for all cells in column 13 or 14 (IR / QFIR)
                cell.style = 'Percent'  # ... change cell format (style) to 'Percent', a built-in style within openpyxl

    wb.save(filename)  # save workbook as admin.xlsx
    logging.debug('Excel workbook completed and saved')


def make_bold(sheet, wb, sheetSlice):
    highlight = NamedStyle(name='highlight')
    highlight.font = Font(bold=True)
    wb.add_named_style(highlight)
    for row in sheetSlice:  # iterate over rows in slice (seems redundant as only 1 row but apparently necessary)
        for cell in row:  # iterate over cells in row
            sheet[cell.coordinate].style = highlight  # add bold to each cell


def export_to_sqlite(listOfProjects): # Export to SQLite
    global conn, c
    logging.debug('Initiating SQLite section')
    conn = sqlite3.connect('admin.db')  # define connection - database is created also
    c = conn.cursor()  # define cursor

    def create_table():
        c.execute(
            'CREATE TABLE IF NOT EXISTS surveysTable(URL TEXT, Alias TEXT, SurveyName TEXT, ProjectNumber TEXT, ClientName TEXT, junk TEXT, ExpectedLOI REAL, ActualLOI REAL, Completes REAL, ScreenOuts REAL, QuotaFulls REAL, LiveOnSite TEXT, IncidenceRate REAL, QFIR REAL)')  # creates the table. CAPS for pure SQL, regular casing otherwise.

    def dynamic_data_entry(
            list):  # at the moment if I pass it an ordered list, it will assign that list to the headings. If I convert dictionariesList into a list of lists, this will be simple.
        # Trying to do a lot on this next line, something is up with it
        c.execute(
            "INSERT INTO surveysTable (URL, Alias, SurveyName, ProjectNumber, ClientName, junk, ExpectedLOI, ActualLOI, Completes, ScreenOuts, QuotaFulls, LiveOnSite, IncidenceRate, QFIR) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (list[0], list[1], list[2], list[3], list[4], list[5], list[6], list[7], list[8], list[9], list[10],
             list[11], list[12], list[13]))
        conn.commit()  # saves to DB. Don't want to close the connection til I'm done using SQL in the program as open/closing wastes resources

    logging.debug('Now calling SQLite fn create_table')
    create_table()  # run the function above

    logging.debug('Now calling SQLite fn dynamic_data_entry')
    for list in listOfProjects:
        dynamic_data_entry(list)  # run the function above

    c.close()
    conn.close()

def send_email(user, pwd, recipient, subject, body):

    gmail_user = user
    gmail_pwd = pwd
    FROM = user
    TO = recipient if type(recipient) is list else [recipient]
    SUBJECT = subject
    TEXT = body

    # Prepare actual message
    message = """From: %s\nTo: %s\nSubject: %s\n\n%s
    """ % (FROM, ", ".join(TO), SUBJECT, TEXT)
    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.ehlo()
        server.starttls()
        server.login(gmail_user, gmail_pwd)
        server.sendmail(FROM, TO, message)
        server.close()
        print('successfully sent the mail')
    except:
        print('failed to send mail')

def email_body_content(listOfNewbies):
    logging.debug('Initialising email_body_content function')
    body = ''
    for projectNumber in listOfNewbies:
        for project in latest10:
            if projectNumber in project:
                #print('Project found:',project)
                body += 'New project added to Admin. Project number: {} ; Project name: {}, Client name: {} \n\n'.format(project[3],project[1],project[4])
    #print(body)
    return(body)

def dictCreator(valueList):   #this function takes in a MO from the regex and creates and returns a per-project dict, with keys as per the headings below
    headings = ['URL','Alias','Survey name','Project number','Client name','junk','Expected LOI','Actual LOI','Completes','Screen Outs','Quota Fulls','Live on site'] #here I've added 'Live on Site'
    newDict = {}
    for i in range(0,len(headings)):
        newDict.setdefault(headings[i], valueList[i])
    completes = int(valueList[8])
    QFs = int(valueList[10])
    SOs = int(valueList[9])
    if completes == 0:
        incidence = 0
        QFincidence = 0
    else:
        try:
            incidence = (completes / (completes + SOs))
        except Exception as err:
            #print ('an exception occured: ', err)
            incidence = 0
        try:
            QFincidence = (completes / (completes + SOs + QFs))
        except Exception as err2:
            # print('an exception occured:',err2)
            QFincidence = 0
    newDict.setdefault('incidence', incidence)
    newDict.setdefault('QFincidence', QFincidence)
    return newDict


def create_masterDict(mo):     #creates a dict of all project dicts in given MO
    mDict = {}
    for i in range(0,len(mo)):
        # logging.debug(f'i is {i}, adding {mo[i][3]} to mDict')
        mDict.setdefault(mo[i][3],dictCreator(mo[i]))
    #TO DO: create a dictionary where each key is the project number and each value is the dict for that job
    #for i in range(0, len(mo) - 1):
    #    mList.append(listCreator(mo[i]))
    return mDict


#AUG-2018 - this is my current working area
#TO DO: create version which runs at 5:30PM and then again at 8:30AM and emails an update of what has changed in that time
#First I want to change my data structure back to dictionaries, so that new and old dictionaries can be combined/compared

moOriginal = process_soup(exampleOldSoup)   #parameter: newSoup or exampleOldSoup for testing
originalDict = create_masterDict(moOriginal)

# now I can create a dictionary of the new content
moNew = process_soup(exampleNewSoup)
latestDict = create_masterDict(moNew)
# excel_export_dict(latestDict, 'admin_dict2.xlsx')

#now I need to create dictionary with old and new data and ultimately show the differences between the two

# Before building the merged dict, I need to create dictionaries which indicate which variable in the old/new data
# dictionaries respectively should be mapped to which variable in the merged dict. e.g. in old data, 'Completes' becomes
# 'Completes_Original'. I've laid this out in excel and will import from there into mapping dictionaries using 'excelToDictConverter'

def excelToDictConverter(excel_filename,r1, r2, c1, c2): # given excel filename and 2-column-wide excel table co-ordinates, creates a dictionary converting the table into key-value pairs
    logging.debug('Now attempting to read-in excel data to create dict')
    map_wb = openpyxl.load_workbook(excel_filename)
    map_sheet = map_wb.active
    dict = {}
    for row in range(r1,r2):
        for column in range(c1,c2):
            map_cell = map_sheet.cell(row = row, column = column)
            v = map_cell.value
            #print(f'row is {row}. Column is {column}, value is {v}')
            if column == c1:
                key = v
            else:
                value = v
        dict.setdefault(key, value)
    return dict

oldMap = excelToDictConverter('mapping.xlsx',3,17,1,3)
newMap = excelToDictConverter('mapping.xlsx',3,17,4,6)

# now I need to create a new dict that contains all the info - new, old and dynamically created, and then export this to excel (perhaps excluding unchanged rows), then have this emailed each morning to KP/JW
# first create dict containing old projects, using modified headings/keys

def createMergedDictWithOldData(oldDataDict, oldDataMappingDict):
    merged = {}
    for k, v in oldDataDict.items():
        nestedDict = {}  # blank dict which we will add to mergedDict at the end of each loop
        for nk, nv in v.items():
            # print(nk, nv)
            equiv = oldDataMappingDict.get(nk)
            if equiv != nk:
                # print(f'project {k} has {nk} re-assigned as {equiv} equal to {nv}')
                nestedDict.setdefault(equiv, nv)
            else:
                # print(f'project {k} has {nk} same as {equiv} so no re-assignment; equal to {nv}')
                nestedDict.setdefault(nk, nv)
        merged.setdefault(k, nestedDict)
    return merged

mergedDict = createMergedDictWithOldData(originalDict, oldMap)

# now add all the new data, bearing in mind that the project may or may not already exist in mergedDict
def addNewData(newDataDict, mergedDataDict, newDataMappingDict):
    for k, v in newDataDict.items():
        nestedDict = {} # blank dict which we will add to mergedDict at the end of each loop
        if k not in mergedDataDict.keys():   # if a totally new project
            for nk, nv in v.items():    # loop through the keys and values of the project
                # print(nk, nv)
                equiv = newDataMappingDict.get(nk)
                nestedDict.setdefault(equiv, nv)
                nestedDict.setdefault('Completes_Original', 0)
                nestedDict.setdefault('Screen Outs_Original', 0)
                nestedDict.setdefault('Quota Fulls_Original', 0)
        else:
            # print(f'{k} found in mergedDict.keys, attempting to add to it')
            for nk, nv in v.items():    # loop through the keys and values of the project
                # print(nk, nv)
                equiv = newDataMappingDict.get(nk)
                if equiv not in mergedDataDict[k].keys():
                    # print(f'adding to {k}: {equiv} = {nv}')
                    mergedDataDict[k][equiv] = nv

        mergedDataDict.setdefault(k, nestedDict)

addNewData(latestDict, mergedDict, newMap)

# now let's add the formula-calculated fields within each dict
def dynamicFieldAdder(dict):  #add the dynamic fields (gaps, overnight) to mergedDict
    for k, v in dict.items():
        c_gap = int(v['Completes_Revised']) - int(v['Completes_Original'])
        v['Completes_gap'] = c_gap
        # print(f'Completes Gap for {k} is {c_gap}')
        s_gap = int(v['Screen Outs_Revised']) - int(v['Screen Outs_Original'])
        v['Screen Outs_gap'] = s_gap
        # print(f'Screen Outs Gap for {k} is {s_gap}')
        q_gap = int(v['Quota Fulls_Revised']) - int(v['Quota Fulls_Original'])
        v['Quota Fulls_gap'] = q_gap
        # print(f'Quota Fulls Gap for {k} is {q_gap}')
        try:
            oIR = (c_gap / (c_gap + s_gap))
            v['incidence_overnight'] = oIR
        except Exception as err:
            #print ('an exception occured: ', err)
            oIR = 0
            v['incidence_overnight'] = oIR
        try:
            oQFIR = (c_gap / (c_gap + s_gap + q_gap))
            v['QFincidence_overnight'] = oQFIR
        except Exception as err:
            #print ('an exception occured: ', err)
            oQFIR = 0
            v['QFincidence_overnight'] = oQFIR

mergedDictHeadings = ['URL',
'Alias',
'Survey name',
'Project number',
'Client name',
'junk',
'Expected LOI',
'Actual LOI',
'Completes_Original',
'Completes_Revised',
'Completes_gap',
'Screen Outs_Original',
'Screen Outs_Revised',
'Screen Outs_gap',
'Quota Fulls_Original',
'Quota Fulls_Revised',
'Quota Fulls_gap',
'Live on site',
'incidence',
'incidence_overnight',
'QFincidence',
'QFincidence_overnight',
]

dynamicFieldAdder(mergedDict) #add the dynamic fields (gaps, overnight) to mergedDict

def excel_export_mergedDict(dict, filename, headings):     #export merged dict to excel
    logging.debug('Attempting to export mergedDict to excel')
    wb = openpyxl.Workbook()  # create excel workbook object
    wb.save(filename)  # save workbook
    sheet = wb.active  # create sheet object as the Active sheet from the workbook object
    wb.save(filename)  # save workbook

    # this bit populates and emboldens the first row
    row = 1
    for column in range(0,len(headings)):
        cell = sheet.cell(row=row, column=column+1)
        cell.value = headings[column]
    make_bold(sheet, wb, sheet['A1':'V1'])    #Calls the make_bold function on first row of excel sheet

    percentageHeadings = ['incidence', 'incidence_overnight', 'QFincidence', 'QFincidence_overnight',]

     #this bit then populates the rest of the sheet with the mergedDict content
    for row, item_tuple in enumerate(dict.items(), 2):
        for column, heading in enumerate(headings, 1):
            cell = sheet.cell(row=row, column=column)  # so on first loop, row = 2, col = 1
            v = item_tuple[1].get(heading)
            try:
                v = float(v)  # try to convert value to a float, so it will store numbers as numbers and not strings
            except ValueError:
                pass  # if it's not a number and therefore returns an error, don't try to convert it to a number
            except TypeError:
                pass
            cell.value = v
            if heading in percentageHeadings:  # for all cells with headings that should have % data
                cell.style = 'Percent'  # ... change cell format (style) to 'Percent', a built-in style within openpyxl
            if heading == 'Completes_gap':
                light_blue = 'A9CCE3'
                cell.fill = PatternFill("solid", fgColor=light_blue)
            if (heading == 'Screen Outs_gap') | (heading == 'Quota Fulls_gap'):
                orange = 'F8C471'
                cell.fill = PatternFill("solid", fgColor=orange)
    wb.save(filename)  # save workbook with given filename
    logging.debug('Excel workbook completed and saved')

excel_export_mergedDict(mergedDict, 'mergedDict.xlsx', mergedDictHeadings) # excel export of mergedDict


# now I need to create a more readable excel export only containing pertinent info / projects

# If Comp, SO or QF gaps > 0, then project has changed. Add it to a 'changed' dictionary, and export that to excel, excluding junk/alias/URL fields


def changesDictCreator(largeDict):
    my_dict = {}
    for k, v in largeDict.items():
        if v['Completes_gap'] > 0 or v['Screen Outs_gap'] > 0 or v['Quota Fulls_gap'] > 0:
            my_dict.setdefault(k, v)
    return my_dict


changesDict = changesDictCreator(mergedDict)

# only certain headings are of interest in the new 'changes' excel export, they are in this list
changesDictHeadingsOfInterest = [
'Survey name','Project number','Client name','Expected LOI','Actual LOI','Completes_Original','Completes_Revised',
    'Completes_gap','Screen Outs_Original','Screen Outs_Revised','Screen Outs_gap','Quota Fulls_Original',
    'Quota Fulls_Revised','Quota Fulls_gap','incidence','incidence_overnight','QFincidence','QFincidence_overnight',
]


excel_export_mergedDict(changesDict, 'changesDict.xlsx', changesDictHeadingsOfInterest) # excel export of changesDict using columns of interest only

# now I've got the report ready I need to work out how I want to store data. Could stash in an xls/database on my home PC rather than have to run a program around the clock.
# looking into standards to see how others would do this - perhaps with a cloud database
# report could have 'last run date/time'





# moNew = process_soup(exampleNewSoup)

'''
### This is where the levers get pulled.

# First we set up the original variables, so this happens outside of the while loop as a one-off

# newSoup = download_soup()     #toggle off for test mode
moOriginal = process_soup(exampleOldSoup)   #parameter: newSoup or exampleOldSoup for testing
#logging.debug('exampleSoup looks like this:\n\n',exampleSoup)
original10 = create_topList(moOriginal, 10)   #match object, desired number of projects in list
while 1:     #this is the loop that endlessly repeats
    #newSoup = download_soup()                # download latest HTML; toggle off for test mode
    mo2 = process_soup(exampleNewSoup)   # parameter can be newSoup for live or exampleNewSoup for test mode
    latest10 = create_topList(mo2, 10)
    newbies = new_project_search(latest10,original10)   #parameters should be latest10 and original10
    print('Latest10 looks like this:\n',latest10)
    print('Original10 looks like this:\n',original10)
    print('newbies:\n',newbies)
    if len(newbies) > 0:
        send_email(cfg.my_gmail_uname, cfg.my_gmail_pw, cfg.my_work_email,'Admin: new project added',email_body_content(newbies))
    original10 = latest10    #overwrite original10 with the latest10
    print('End of program, waiting 60 sec')
    time.sleep(1000)     #1000 for test mode



'''







"""
original20 = create_topList(moOriginal, 20)   #match object, desired number of projects in list
new20 = create_topList(moNew, 20)
print("new projects are: ", new_project_search(new20, original20))
print("original20 looks like this: ",original20)
print("the job# in the first item in original20 looks like this: ", original20[0][3])
"""


# for all job #s in new20, if job# appears in original20:
    # for all non-numerical items, their equivalent in changed20 is identical to new20
    # for all numerical items in that job for new20:
        # changed20 equivalent = new20 number minus original20 number







#TO DO: compare original10 and latest10 and flag any 'zero to 1' completes movement(new function)


#masterList = create_masterList(mo2)   #match object
#excel_export(latest10)           #list to export
#export_to_sqlite(original10)       #list to export



#This is a test sequence, to compare lists generated from old and new soup
#It works beautifully when I'm looking for 10 and 20 list length, but for 30 I get an error. Not sure why a project was being searched for and attempted removal twice, but added 'try and except' logic to keep program running


# logging.debug('Example sequence')
# exampleOldMo = process_soup(exampleOldSoup)
# exampleNewMo = process_soup(exampleNewSoup)
# exampleOriginal10 = create_topList(exampleOldMo, 10)   #match object, desired number of projects in list
# print('ExampleOriginal10:\n', exampleOriginal10,'\n')
# exampleLatest10 = create_topList(exampleNewMo, 10)   #match object, desired number of projects in list
# print('ExampleLatest10:\n', exampleLatest10,'\n')
# exampleNewbies = new_project_search(exampleLatest10, exampleOriginal10)
# print('The example new projects are:\n',exampleNewbies,'\n')
# send_email(cfg.my_gmail_uname, cfg.my_gmail_pw, cfg.my_work_email,'Admin: new project added',str(exampleNewbies))

